"""
Load GymAware athlete allowlist from the client's reference workbook.

Default path: GymAware API Reference Numbers.xlsx in project root, or GYMAWARE_ALLOWLIST_XLSX.
Use for Option A privacy filtering (summaries/reps → only rows whose athleteReference is in the set).
"""
from __future__ import annotations

import os
from pathlib import Path
from typing import Any

DEFAULT_REL = "GymAware API Reference Numbers.xlsx"


def _project_root() -> Path:
    return Path(__file__).resolve().parents[2]


def load_athlete_references_from_xlsx(
    path: str | Path | None = None,
) -> tuple[list[dict[str, Any]], set[int]]:
    """
    Parse Sheet1 with columns: Last Name, First Name, GymAware API ID (4th column in file).

    Returns (rows_as_dicts, reference_id_set).
    """
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("Install openpyxl: pip install openpyxl") from e

    if path is None:
        path = os.getenv("GYMAWARE_ALLOWLIST_XLSX", "").strip()
    if not path:
        path = _project_root() / DEFAULT_REL
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(f"Allowlist workbook not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_out: list[dict[str, Any]] = []
        refs: set[int] = set()
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row or len(row) < 4:
                continue
            last = row[1]
            first = row[2]
            ref = row[3]
            if ref is None or str(ref).strip() == "":
                continue
            if isinstance(ref, str) and "API ID" in ref:
                continue  # header
            try:
                ref_int = int(ref)
            except (TypeError, ValueError):
                continue
            ln = str(last).strip() if last is not None else ""
            fn = str(first).strip() if first is not None else ""
            rows_out.append(
                {
                    "last_name": ln,
                    "first_name": fn,
                    "athlete_reference": ref_int,
                }
            )
            refs.add(ref_int)
        return rows_out, refs
    finally:
        wb.close()


def athlete_reference_allowlist() -> set[int]:
    """Convenience: only the GymAware athleteReference IDs."""
    _, refs = load_athlete_references_from_xlsx()
    return refs


def env_use_allowlist() -> bool:
    """True when GYMAWARE_USE_ALLOWLIST is 1/true/yes/on (export + upload respect this)."""
    v = os.getenv("GYMAWARE_USE_ALLOWLIST", "").strip().lower()
    return v in ("1", "true", "yes", "on")


def filter_rows_by_athlete_reference(
    rows: list[dict[str, Any]],
    allow: set[int],
) -> list[dict[str, Any]]:
    """Keep API rows whose athleteReference is in allow (GymAware camelCase)."""
    out: list[dict[str, Any]] = []
    for row in rows:
        ar = row.get("athleteReference")
        try:
            ar_int = int(ar) if ar is not None else None
        except (TypeError, ValueError):
            continue
        if ar_int is not None and ar_int in allow:
            out.append(row)
    return out
